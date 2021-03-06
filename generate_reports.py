from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation
from pony.orm import db_session
from setuptools.namespaces import flatten

from boomer_utils import serialize_yes_or_no, adjust_cell_sizes, adjust_cell_sizes_for_judge_feedback
from models import Event, School, Participant

MASTER_REPORT = "output/Master.Report.xlsx"
JUDGE_REPORT = "output/Judge.Report.docx"


@db_session
def generate_master_report():
    print("Generating master report")

    events = Event.select().order_by(Event.name)
    schools = School.select().order_by(School.name)
    workbook = Workbook()
    event_worksheet = workbook.active
    event_worksheet.title = 'Events'
    for event in events:
        event_worksheet.append([event.name, len(event.registrations)])
    adjust_cell_sizes(event_worksheet)

    school_worksheet = workbook.create_sheet('Schools')
    school_worksheet.append([
        "School",
        "Regular Registrations",
        "Late Registrations",
        "Registration Fees",
        "Total Enrolled",
        "Rookie Teacher",
        "Rookie School",
        "Attending State"
    ])
    for school in schools:
        school_worksheet.append([
            school.name,
            school.regular_registrations,
            school.late_registrations,
            school.regular_registrations * 10 + school.late_registrations * 12,
            school.total_enrolled,
            serialize_yes_or_no(school.rookie_teacher),
            serialize_yes_or_no(school.rookie_school),
            serialize_yes_or_no(school.attending_state)
        ])
    adjust_cell_sizes(school_worksheet)

    workbook.save(MASTER_REPORT)


@db_session
def generate_event_sheets():
    events = Event.select().order_by(Event.name)
    for event in events:
        generate_event_sheet(event)


def generate_event_sheet(event):
    if len(event.registrations) > 0:
        print("Generating", event.name, "judge sheet")
    else:
        print("No registrations for", event.name)
        return
    workbook = Workbook()
    worksheet = workbook.active

    # Fill headers
    worksheet.append([
        "Participant(s)",
        "School",
        "Score on Crit. 1, Judge 1",
        "Score on Crit. 2, Judge 1",
        "Score on Crit. 3, Judge 1",
        "Score on Crit. 4, Judge 1",
        "Score on Crit. 5, Judge 1",
        "Total, Judge 1",
        "Comments, Judge 1",
        "Score on Crit. 1, Judge 2",
        "Score on Crit. 2, Judge 2",
        "Score on Crit. 3, Judge 2",
        "Score on Crit. 4, Judge 2",
        "Score on Crit. 5, Judge 2",
        "Total, Judge 2",
        "Comments, Judge 2",
        "Score on Crit. 1, Judge 3",
        "Score on Crit. 2, Judge 3",
        "Score on Crit. 3, Judge 3",
        "Score on Crit. 4, Judge 3",
        "Score on Crit. 5, Judge 3",
        "Total, Judge 3",
        "Comments, Judge 3",
        "One-time deduction, if any (enter as a positive number)",
        "Total Score"
    ])

    # Set formulas
    for row, registration in enumerate(event.registrations, 2):
        worksheet.append([
            '\n'.join(p.name for p in registration.participants.order_by(Participant.name)),
            registration.school.name,
            '', '', '', '', '',
            f"=SUM(C{row}:G{row})",  # Total judge 1
            '', '', '', '', '', '',
            f"=SUM(J{row}:N{row})",  # Total judge 2
            '', '', '', '', '', '',
            f"=SUM(Q{row}:U{row})",  # Total judge 3
            '', '',
            f"=H{row} + O{row} + V{row} - X{row}"  # Total score
        ])

    # Lock worksheet while unlocking judge input cells
    worksheet.protection.sheet = True
    max_row = len(event.registrations) + 1
    input_cell_panes = worksheet[f"C2:G{max_row}"] \
        + worksheet[f"I2:N{max_row}"] \
        + worksheet[f"P2:U{max_row}"] \
        + worksheet[f"W2:X{max_row}"]
    input_cells = flatten(input_cell_panes)
    for cell in input_cells:
        cell.protection = Protection(locked=False)

    # Only allow point values from zero to twenty
    point_validator = DataValidation(type='whole', operator='between', formula1=0, formula2=20)
    worksheet.add_data_validation(point_validator)
    point_validator.add(f"C2:G{max_row}")
    point_validator.add(f"J2:N{max_row}")
    point_validator.add(f"Q2:U{max_row}")
    point_validator.add(f"X2:X{max_row}")

    # Finalize and save
    adjust_cell_sizes_for_judge_feedback(worksheet)
    event_sheet = F"output/Event.{event.name}.xlsx"
    workbook.save(event_sheet)
